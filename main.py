"""
Markirovka Tekshirish Tizimi — Android (Kivy)
"""
import json, os, sys, wave, struct, math, threading
from datetime import datetime

import kivy
kivy.require('2.3.0')

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.progressbar import ProgressBar
from kivy.uix.widget import Widget
from kivy.uix.tabbedpanel import TabbedPanel, TabbedPanelItem
from kivy.uix.filechooser import FileChooserListView
from kivy.core.window import Window
from kivy.utils import get_color_from_hex
from kivy.metrics import dp, sp
from kivy.clock import Clock
from kivy.graphics import Color, Rectangle, RoundedRectangle

# Android-specific
try:
    from android.storage import primary_external_storage_path
    from android.permissions import request_permissions, Permission
    ANDROID = True
except ImportError:
    ANDROID = False

# ── Ranglar ──────────────────────────────────────────────────────────────
C_DARK   = get_color_from_hex('#1A2B4A')
C_BLUE   = get_color_from_hex('#2563EB')
C_GREEN  = get_color_from_hex('#059669')
C_RED    = get_color_from_hex('#DC2626')
C_AMBER  = get_color_from_hex('#D97706')
C_OK_BG  = get_color_from_hex('#D1FAE5')
C_OK_FG  = get_color_from_hex('#065F46')
C_ERR_BG = get_color_from_hex('#FEE2E2')
C_ERR_FG = get_color_from_hex('#991B1B')
C_DUP_BG = get_color_from_hex('#FEF3C7')
C_DUP_FG = get_color_from_hex('#92400E')
C_BG     = get_color_from_hex('#F8FAFC')
C_WHITE  = get_color_from_hex('#FFFFFF')
C_BORDER = get_color_from_hex('#E2E8F0')
C_MUTED  = get_color_from_hex('#64748B')
C_PANEL  = get_color_from_hex('#F1F5F9')


def _xml(text):
    if not isinstance(text, str):
        text = str(text)
    return (text.replace('&','&amp;').replace('<','&lt;')
                .replace('>','&gt;').replace('"','&quot;')
                .replace("'","&#39;"))


def _get_storage():
    if ANDROID:
        return primary_external_storage_path()
    return os.path.expanduser('~')


def _get_sounds_dir():
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, 'sounds')


def _generate_sound(filepath, tones):
    SR = 44100
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with wave.open(filepath, 'w') as wf:
        wf.setnchannels(1); wf.setsampwidth(2); wf.setframerate(SR)
        frames = []
        for freq, dur in tones:
            n = int(SR * dur / 1000)
            for i in range(n):
                if freq == 0:
                    frames.append(struct.pack('<h', 0))
                else:
                    t = i / SR
                    fi = int(SR*0.005); fo = int(SR*0.04)
                    env = (i/fi if i<fi else (n-i)/fo if i>n-fo else 1.0)
                    val = math.sin(2*math.pi*freq*t)
                    frames.append(struct.pack('<h', int(val*env*0.7*32767)))
        wf.writeframes(b''.join(frames))


def _ensure_sounds():
    sd = _get_sounds_dir()
    defs = {
        'ok.wav':  [(880,100),(0,30),(1320,130)],
        'err.wav': [(400,90),(0,25),(260,220)],
        'dup.wav': [(660,70),(0,40),(660,70),(0,40),(660,70)],
    }
    for fname, tones in defs.items():
        fp = os.path.join(sd, fname)
        if not os.path.exists(fp):
            try: _generate_sound(fp, tones)
            except: pass


def play_sound(name):
    def _play():
        try:
            fp = os.path.join(_get_sounds_dir(), f'{name}.wav')
            if not os.path.exists(fp):
                _ensure_sounds()
            if ANDROID:
                try:
                    from kivy.core.audio import SoundLoader
                    s = SoundLoader.load(fp)
                    if s: s.play()
                except: pass
            elif sys.platform == 'win32':
                import winsound
                winsound.PlaySound(fp, winsound.SND_FILENAME|winsound.SND_ASYNC)
            else:
                import subprocess
                cmd = ['afplay',fp] if sys.platform=='darwin' else ['aplay','-q',fp]
                subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except: pass
    threading.Thread(target=_play, daemon=True).start()

threading.Thread(target=_ensure_sounds, daemon=True).start()


# ── HTML hisobot (PDF o'rniga) ───────────────────────────────────────────
def _make_pdf(path, products, scanned_set, scan_log, stats, fi):
    """PDF o'rniga HTML hisobot — Android da reportlab ishlamaydi."""
    report_date = datetime.now().strftime('%Y-%m-%d %H:%M')
    total_marks = sum(len(p['marks']) for p in products)
    missing = total_marks - stats['ok']

    # HTML faylga saqlash (.html kengaytmasi bilan)
    html_path = path.replace('.pdf', '.html')

    ok_pct = int(len(scanned_set) / total_marks * 100) if total_marks > 0 else 0

    rows_products = ''
    for p in products:
        sc = sum(1 for m in p['marks'] if m in scanned_set)
        mc = len(p['marks']) - sc
        color = '#D1FAE5' if mc == 0 else '#FEE2E2'
        txt   = '#065F46' if mc == 0 else '#991B1B'
        rows_products += (
            f'<tr style="background:{color}">'
            f'<td style="padding:6px 8px;color:{txt}">{p["ordno"]}</td>'
            f'<td style="padding:6px 8px;color:{txt}">{_xml(p["name"][:60])}</td>'
            f'<td style="padding:6px 8px;text-align:center;color:{txt}">{len(p["marks"])}</td>'
            f'<td style="padding:6px 8px;text-align:center;color:{txt}">{sc}</td>'
            f'<td style="padding:6px 8px;text-align:center;color:{txt}">{mc}</td>'
            f'</tr>'
        )

    rows_missing = ''
    cnt = 0
    for p in products:
        for m in p['marks']:
            if m not in scanned_set:
                cnt += 1
                bg = '#FEF2F2' if cnt % 2 == 0 else '#FFFFFF'
                rows_missing += (
                    f'<tr style="background:{bg}">'
                    f'<td style="padding:6px 8px">{cnt}</td>'
                    f'<td style="padding:6px 8px;font-family:monospace;font-size:12px;color:#991B1B">{_xml(m)}</td>'
                    f'<td style="padding:6px 8px">{_xml(p["name"][:50])}</td>'
                    f'</tr>'
                )

    rows_log = ''
    for i, r in enumerate(scan_log[:100]):
        st = r['status']
        bg = '#D1FAE5' if st=='ok' else ('#FEF3C7' if st=='dup' else '#FEE2E2')
        lbl = "To\'g\'ri" if st=='ok' else ('Takroriy' if st=='dup' else 'Xato')
        rows_log += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:6px 8px">{lbl}</td>'
            f'<td style="padding:6px 8px;font-family:monospace;font-size:11px">{_xml(r["mark"][:35])}</td>'
            f'<td style="padding:6px 8px">{_xml(r.get("prodname","")[:40])}</td>'
            f'</tr>'
        )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Markirovka Hisoboti - Faktura {_xml(fi.get('no',''))}</title>
<style>
body{{font-family:Arial,sans-serif;margin:0;padding:16px;background:#F8FAFC;color:#1E293B}}
h1{{background:#1A2B4A;color:#fff;padding:16px;margin:-16px -16px 16px;font-size:18px}}
h2{{font-size:15px;color:#1A2B4A;margin:20px 0 8px}}
.info{{background:#fff;border:1px solid #E2E8F0;border-radius:8px;padding:12px;margin-bottom:12px}}
.info td{{padding:4px 8px;font-size:13px}}
.info td:first-child{{color:#64748B;font-weight:bold;white-space:nowrap}}
.stats{{display:grid;grid-template-columns:repeat(2,1fr);gap:8px;margin-bottom:16px}}
.stat{{background:#fff;border:1px solid #E2E8F0;border-radius:8px;padding:12px;text-align:center}}
.stat .num{{font-size:28px;font-weight:bold}}
.stat .lbl{{font-size:12px;color:#64748B}}
.prog{{background:#E2E8F0;border-radius:4px;height:12px;margin:8px 0}}
.prog-fill{{height:12px;border-radius:4px;background:#059669}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;margin-bottom:16px}}
th{{background:#2563EB;color:#fff;padding:8px;font-size:12px;text-align:left}}
td{{font-size:12px;border-bottom:1px solid #E2E8F0}}
.miss{{background:#FEE2E2;color:#991B1B;padding:10px 14px;border-radius:8px;margin-bottom:12px;font-weight:bold}}
.ok{{background:#D1FAE5;color:#065F46;padding:10px 14px;border-radius:8px;margin-bottom:12px;font-weight:bold}}
.footer{{font-size:11px;color:#94A3B8;text-align:center;margin-top:20px}}
</style></head><body>
<h1>Markirovka Tekshirish Hisoboti</h1>
<div class="info"><table>
<tr><td>Faktura:</td><td>&#8470;{_xml(fi.get('no',''))} &nbsp; {_xml(fi.get('date',''))}</td></tr>
<tr><td>Sotuvchi:</td><td>{_xml(fi.get('seller_name',''))}</td></tr>
<tr><td>TIN:</td><td>{_xml(fi.get('seller_tin',''))}</td></tr>
<tr><td>Rahbar:</td><td>{_xml(fi.get('seller_director',''))}</td></tr>
<tr><td>Xaridor:</td><td>{_xml(fi.get('buyer_name',''))}</td></tr>
<tr><td>TIN:</td><td>{_xml(fi.get('buyer_tin',''))}</td></tr>
<tr><td>Rahbar:</td><td>{_xml(fi.get('buyer_director',''))}</td></tr>
</table></div>
<div class="stats">
<div class="stat"><div class="num" style="color:#2563EB">{stats['total']}</div><div class="lbl">Jami skanerlangan</div></div>
<div class="stat"><div class="num" style="color:#059669">{stats['ok']}</div><div class="lbl">To'g'ri</div></div>
<div class="stat"><div class="num" style="color:#DC2626">{stats['err']}</div><div class="lbl">Xato</div></div>
<div class="stat"><div class="num" style="color:#D97706">{stats['dup']}</div><div class="lbl">Takroriy</div></div>
</div>
<div class="prog"><div class="prog-fill" style="width:{ok_pct}%"></div></div>
<p style="text-align:center;font-size:13px;color:#64748B">{len(scanned_set)} / {total_marks} skanerlangan ({ok_pct}%)</p>
{'<div class="miss">&#9888; ' + str(missing) + ' ta markirovka hali skanerlanmagan!</div>' if missing > 0 else '<div class="ok">&#10003; Barcha markirovkalar skanerlangan!</div>'}
<h2>Mahsulotlar</h2>
<table><tr><th>#</th><th>Mahsulot</th><th>Jami</th><th>Skaner</th><th>Qoldi</th></tr>
{rows_products}</table>
{'<h2>Yetishmayotgan markirovkalar</h2><table><tr><th>#</th><th>Markirovka kodi</th><th>Mahsulot</th></tr>' + rows_missing + '</table>' if missing > 0 else ''}
<h2>Skanerlash jurnali</h2>
<table><tr><th>Natija</th><th>Markirovka kodi</th><th>Mahsulot</th></tr>
{rows_log}</table>
<div class="footer">Hisobot: {report_date} | Faktura &#8470;{_xml(fi.get('no',''))} | Jami {total_marks} ta markirovka</div>
</body></html>"""

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return html_path


# ── Excel generator ───────────────────────────────────────────────────────
def _make_excel(path, products, scanned_set, scan_log, stats, fi):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    report_date = datetime.now().strftime('%Y-%m-%d %H:%M')
    total_marks = sum(len(p['marks']) for p in products)
    missing = total_marks - stats['ok']

    C_DARK='1A2B4A'; C_HEAD='2563EB'
    C_OK='D1FAE5'; C_OK_T='065F46'
    C_ERR='FEE2E2'; C_ERR_T='991B1B'
    C_DUP='FEF3C7'; C_DUP_T='92400E'
    C_MISS='FEF2F2'; C_ALT='F1F5F9'; C_WHITE='FFFFFF'
    C_BORDER='CBD5E1'; C_INFO='F8FAFC'

    thin = Side(style='thin', color=C_BORDER)
    bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)
    def fill(h): return PatternFill('solid',fgColor=h)
    def hf(sz=10,bold=True,color='FFFFFF'):
        return Font(name='Arial',size=sz,bold=bold,color=color)
    def cf(sz=9,bold=False,color='1E293B'):
        return Font(name='Arial',size=sz,bold=bold,color=color)
    def ac(h='center',v='center'): return Alignment(horizontal=h,vertical=v)

    ws = wb.active; ws.title='Umumiy hisobot'
    ws.sheet_view.showGridLines=False
    for col,w in {'A':3,'B':22,'C':10,'D':35,'E':10,'F':12,'G':12,'H':12,'I':3}.items():
        ws.column_dimensions[col].width=w

    ws.row_dimensions[1].height=6; ws.row_dimensions[2].height=38
    ws.merge_cells('B2:H2')
    t=ws['B2']; t.value='MARKIROVKA TEKSHIRISH HISOBOTI'
    t.font=Font(name='Arial',size=17,bold=True,color='FFFFFF')
    t.fill=fill(C_DARK); t.alignment=ac()

    ws.row_dimensions[3].height=6
    meta=[
        (4,18,'Faktura raqami:',f"№{fi.get('no','')}","Faktura sanasi:",fi.get('date','')),
        (5,18,'Sotuvchi TIN:',fi.get('seller_tin',''),'Rahbar:',fi.get('seller_director','')),
        (6,18,'Sotuvchi:',fi.get('seller_name',''),'Manzil:',fi.get('seller_address','')),
        (7,18,'Xaridor TIN:',fi.get('buyer_tin',''),'Rahbar:',fi.get('buyer_director','')),
        (8,18,'Xaridor:',fi.get('buyer_name',''),'Manzil:',fi.get('buyer_address','')),
    ]
    for row,rh,l1,v1,l2,v2 in meta:
        ws.row_dimensions[row].height=rh
        ws.merge_cells(f'B{row}:C{row}'); ws.merge_cells(f'D{row}:E{row}')
        ws.merge_cells(f'F{row}:G{row}')
        for col,val,bold,bg in [('B',l1,True,C_INFO),('D',v1,False,C_WHITE),
                                  ('F',l2,True,C_INFO),('H',v2,False,C_WHITE)]:
            c=ws[f'{col}{row}']; c.value=val; c.fill=fill(bg)
            c.font=Font(name='Arial',size=9,bold=bold,color='64748B' if bold else '1E293B')
            c.alignment=Alignment(vertical='center',wrap_text=True); c.border=bdr

    ws.row_dimensions[9].height=8; ws.row_dimensions[10].height=12
    ws.row_dimensions[11].height=30; ws.row_dimensions[12].height=18
    ws['B10'].value='Umumiy statistika'
    ws['B10'].font=Font(name='Arial',size=9,bold=True,color='64748B')
    for col,lbl,val,tc,bg in [
        ('E','Jami skanerlangan',stats['total'],'2563EB','EFF6FF'),
        ('F',"To'g'ri",stats['ok'],'059669','ECFDF5'),
        ('G','Xato',stats['err'],'DC2626','FEF2F2'),
        ('H','Takroriy',stats['dup'],'D97706','FFFBEB')]:
        cv=ws[f'{col}11']; cv.value=val
        cv.font=Font(name='Arial',size=20,bold=True,color=tc)
        cv.fill=fill(bg); cv.alignment=ac(); cv.border=bdr
        cl=ws[f'{col}12']; cl.value=lbl
        cl.font=Font(name='Arial',size=8,color=tc)
        cl.fill=fill(bg); cl.alignment=ac(); cl.border=bdr

    ws.row_dimensions[13].height=8; ws.row_dimensions[14].height=20
    ws.merge_cells('B14:H14'); mc=ws['B14']
    mc.value=(f'Yetishmayotgan: {missing} ta | Jami: {total_marks} ta | Hisobot: {report_date}')
    mc.font=Font(name='Arial',size=9,bold=True,color=C_ERR_T if missing>0 else C_OK_T)
    mc.fill=fill(C_ERR if missing>0 else C_OK); mc.alignment=ac(); mc.border=bdr

    ws.row_dimensions[15].height=8; ws.row_dimensions[16].height=20
    for col,lbl in [('B','#'),('C','Tartib'),('D','Mahsulot nomi'),
                     ('E','Jami'),('F','Skanerlangan'),('G','Yetishmaydi'),('H','Holat')]:
        c=ws[f'{col}16']; c.value=lbl; c.font=hf(9)
        c.fill=fill(C_HEAD); c.alignment=ac(); c.border=bdr

    row=17
    for i,p in enumerate(products):
        ws.row_dimensions[row].height=17
        sc=sum(1 for m in p['marks'] if m in scanned_set)
        mc2=len(p['marks'])-sc; bg4=C_ALT if i%2==0 else C_WHITE
        st_t="To'liq" if mc2==0 else f'{mc2} ta yetishmaydi'
        st_b=C_OK if mc2==0 else C_ERR; st_f=C_OK_T if mc2==0 else C_ERR_T
        for col,val,align in [('B',p['ordno'],'center'),('C',f"#{p['ordno']}",'center'),
                                ('D',p['name'],'left'),('E',len(p['marks']),'center'),
                                ('F',sc,'center'),('G',mc2,'center'),('H',st_t,'center')]:
            c=ws[f'{col}{row}']; c.value=val; c.border=bdr
            c.alignment=Alignment(horizontal=align,vertical='center',wrap_text=(col=='D'))
            if col=='H':
                c.font=Font(name='Arial',size=9,bold=True,color=st_f); c.fill=fill(st_b)
            else:
                c.font=cf(); c.fill=fill(bg4)
        row+=1

    ws.row_dimensions[row].height=20
    for col in ['B','C','D','E','F','G','H']:
        c=ws[f'{col}{row}']; c.font=hf(9); c.fill=fill(C_DARK); c.border=bdr; c.alignment=ac()
    ws[f'B{row}'].value='JAMI'
    for col,rng in [('E',f'E17:E{row-1}'),('F',f'F17:F{row-1}'),('G',f'G17:G{row-1}')]:
        ws[f'{col}{row}'].value=f'=SUM({rng})'

    ws2=wb.create_sheet('Skanerlash jurnali')
    ws2.sheet_view.showGridLines=False
    for col,w in {'A':3,'B':6,'C':44,'D':40,'E':12,'F':3}.items():
        ws2.column_dimensions[col].width=w
    ws2.row_dimensions[1].height=6; ws2.row_dimensions[2].height=30
    ws2.merge_cells('B2:E2')
    t2=ws2['B2']; t2.value=f"Skanerlash jurnali — Faktura №{fi.get('no','')} ({fi.get('date','')})"
    t2.font=Font(name='Arial',size=13,bold=True,color='FFFFFF')
    t2.fill=fill(C_DARK); t2.alignment=ac()
    ws2.row_dimensions[3].height=8; ws2.row_dimensions[4].height=20
    for col,lbl in [('B','#'),('C','Markirovka kodi'),('D','Mahsulot nomi'),('E','Natija')]:
        c=ws2[f'{col}4']; c.value=lbl; c.font=hf(9)
        c.fill=fill(C_HEAD); c.alignment=ac(); c.border=bdr
    labels={'ok':"To'g'ri",'err':'Xato','dup':'Takroriy'}
    for i,r in enumerate(scan_log,1):
        rr=i+4; ws2.row_dimensions[rr].height=16
        bg2=C_ALT if i%2==0 else C_WHITE; st=r['status']
        st_t=labels[st]; st_b=C_OK if st=='ok' else(C_DUP if st=='dup' else C_ERR)
        st_f=C_OK_T if st=='ok' else(C_DUP_T if st=='dup' else C_ERR_T)
        for col,val,align in [('B',i,'center'),('C',r['mark'],'left'),
                                ('D',r.get('prodname','')[:55],'left'),('E',st_t,'center')]:
            c=ws2[f'{col}{rr}']; c.value=val; c.border=bdr
            c.alignment=Alignment(horizontal=align,vertical='center')
            if col=='E':
                c.font=Font(name='Arial',size=9,bold=True,color=st_f); c.fill=fill(st_b)
            else:
                c.font=cf(); c.fill=fill(bg2)

    ws3=wb.create_sheet('Yetishmayotganlar')
    ws3.sheet_view.showGridLines=False
    for col,w in {'A':3,'B':6,'C':44,'D':40,'E':3}.items():
        ws3.column_dimensions[col].width=w
    ws3.row_dimensions[1].height=6; ws3.row_dimensions[2].height=30
    ws3.merge_cells('B2:D2')
    t3=ws3['B2']; t3.value=f"Yetishmayotgan markirovkalar — Faktura №{fi.get('no','')}"
    t3.font=Font(name='Arial',size=13,bold=True,color='FFFFFF')
    t3.fill=fill(C_DARK); t3.alignment=ac()
    ws3.row_dimensions[3].height=8; ws3.row_dimensions[4].height=20
    for col,lbl in [('B','#'),('C','Markirovka kodi'),('D','Mahsulot nomi')]:
        c=ws3[f'{col}4']; c.value=lbl; c.font=hf(9)
        c.fill=fill(C_HEAD); c.alignment=ac(); c.border=bdr
    miss_row=5; cnt=0
    for p in products:
        for m in p['marks']:
            if m not in scanned_set:
                cnt+=1; ws3.row_dimensions[miss_row].height=16
                bg3=C_MISS if cnt%2==0 else C_WHITE
                for col,val,align in [('B',cnt,'center'),('C',m,'left'),('D',p['name'][:55],'left')]:
                    c=ws3[f'{col}{miss_row}']; c.value=val
                    c.font=cf(); c.fill=fill(bg3); c.border=bdr
                    c.alignment=Alignment(horizontal=align,vertical='center')
                miss_row+=1
    if cnt==0:
        ws3.merge_cells('B5:D5'); c=ws3['B5']
        c.value='Barcha markirovkalar skanerlangan!'
        c.font=Font(name='Arial',size=10,bold=True,color=C_OK_T)
        c.fill=fill(C_OK); c.alignment=ac(); c.border=bdr

    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════
#  UI YORDAMCHI WIDGETLAR
# ══════════════════════════════════════════════════════════════════════════
class RoundedBox(BoxLayout):
    def __init__(self, bg_color=None, radius=10, **kw):
        super().__init__(**kw)
        self._bg = bg_color or C_PANEL
        self._r  = radius
        self.bind(pos=self._redraw, size=self._redraw)

    def _redraw(self, *_):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self._bg)
            RoundedRectangle(pos=self.pos, size=self.size,
                             radius=[self._r])


class StatCard(BoxLayout):
    def __init__(self, label, value_var, fg, bg, **kw):
        super().__init__(orientation='vertical', padding=dp(8),
                         spacing=dp(2), **kw)
        self._bg = bg
        self.bind(pos=self._draw, size=self._draw)
        self.val_lbl = Label(text='0', font_size=sp(28),
                             bold=True, color=fg,
                             size_hint_y=None, height=dp(40))
        self.txt_lbl = Label(text=label, font_size=sp(11),
                             color=fg, size_hint_y=None, height=dp(20))
        self.add_widget(self.val_lbl)
        self.add_widget(self.txt_lbl)

    def _draw(self, *_):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self._bg)
            RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(8)])

    def set_value(self, v):
        self.val_lbl.text = str(v)


class KivyBtn(Button):
    def __init__(self, bg=None, fg=None, **kw):
        super().__init__(**kw)
        self._bg = bg or C_BLUE
        self._fg = fg or C_WHITE
        self.background_color = (0,0,0,0)
        self.color = self._fg
        self.font_size = sp(14)
        self.bold = True
        self.bind(pos=self._draw, size=self._draw)

    def _draw(self, *_):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self._bg)
            RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(8)])


# ══════════════════════════════════════════════════════════════════════════
#  ASOSIY EKRAN
# ══════════════════════════════════════════════════════════════════════════
class MainScreen(BoxLayout):
    def __init__(self, **kw):
        super().__init__(orientation='vertical', spacing=0, **kw)

        self.products    = []
        self.all_marks   = {}
        self.scanned_set = set()
        self.scan_log    = []
        self.stats       = {'total':0,'ok':0,'err':0,'dup':0}
        self.fi          = {}

        with self.canvas.before:
            Color(*C_BG)
            self._bg_rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=self._upd_bg, size=self._upd_bg)

        self._build_header()
        self._build_body()

    def _upd_bg(self, *_):
        self._bg_rect.pos  = self.pos
        self._bg_rect.size = self.size

    # ── Header ────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = BoxLayout(orientation='horizontal', size_hint_y=None,
                        height=dp(56), padding=[dp(14),dp(8)])
        with hdr.canvas.before:
            Color(*C_DARK)
            self._hdr_rect = Rectangle(pos=hdr.pos, size=hdr.size)
        hdr.bind(pos=lambda *_: setattr(self._hdr_rect,'pos',hdr.pos),
                 size=lambda *_: setattr(self._hdr_rect,'size',hdr.size))

        self.lbl_title = Label(text='Markirovka Tizimi',
                               font_size=sp(16), bold=True,
                               color=C_WHITE, halign='left',
                               size_hint_x=0.6)
        self.lbl_title.bind(size=lambda w,_: setattr(w,'text_size',w.size))

        btn_open = KivyBtn(text='JSON ochish', bg=get_color_from_hex('#334155'),
                           size_hint_x=0.4, height=dp(38), size_hint_y=None)
        btn_open.bind(on_press=self._open_json)

        hdr.add_widget(self.lbl_title)
        hdr.add_widget(btn_open)
        self.add_widget(hdr)

    # ── Body (TabbedPanel) ────────────────────────────────────────────────
    def _build_body(self):
        tp = TabbedPanel(do_default_tab=False)
        tp.tab_width = dp(120)

        # Tab 1: Skanerlash
        t1 = TabbedPanelItem(text='Skanerlash')
        t1.add_widget(self._build_scan_tab())
        tp.add_widget(t1)
        tp.default_tab = t1

        # Tab 2: Jurnal
        t2 = TabbedPanelItem(text='Jurnal')
        t2.add_widget(self._build_log_tab())
        tp.add_widget(t2)

        # Tab 3: Yetishmayotgan
        t3 = TabbedPanelItem(text='Qolganlar')
        t3.add_widget(self._build_missing_tab())
        tp.add_widget(t3)

        self.add_widget(tp)

    # ── Skanerlash tab ────────────────────────────────────────────────────
    def _build_scan_tab(self):
        root = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))

        # Statistika kartalar
        stat_grid = GridLayout(cols=2, spacing=dp(6),
                               size_hint_y=None, height=dp(110))
        self.card_total = StatCard('Jami skanerlandi', None, C_BLUE,   get_color_from_hex('#EFF6FF'))
        self.card_ok    = StatCard("To'g'ri",          None, C_GREEN,  C_OK_BG)
        self.card_err   = StatCard('Xato',             None, C_RED,    C_ERR_BG)
        self.card_dup   = StatCard('Takroriy',         None, C_AMBER,  C_DUP_BG)
        for card in [self.card_total, self.card_ok, self.card_err, self.card_dup]:
            stat_grid.add_widget(card)
        root.add_widget(stat_grid)

        # Progress blok
        prog_box = BoxLayout(orientation='vertical', spacing=dp(4),
                             size_hint_y=None, height=dp(80))

        prog_top = BoxLayout(orientation='horizontal', spacing=dp(6),
                             size_hint_y=None, height=dp(50))
        self.card_scanned  = StatCard('Skanerlandi', None, C_GREEN, C_OK_BG)
        self.card_remain   = StatCard('Qoldi',       None, C_RED,   C_ERR_BG)
        self.card_total_m  = StatCard('Jami (fatura)',None,C_BLUE,  get_color_from_hex('#EFF6FF'))
        for c in [self.card_scanned, self.card_remain, self.card_total_m]:
            prog_top.add_widget(c)
        prog_box.add_widget(prog_top)

        pb_row = BoxLayout(orientation='horizontal', spacing=dp(6),
                           size_hint_y=None, height=dp(26))
        self.prog_bar = ProgressBar(max=100, value=0)
        self.lbl_pct  = Label(text='0%', font_size=sp(12), bold=True,
                              color=C_DARK, size_hint_x=None, width=dp(40))
        pb_row.add_widget(self.prog_bar)
        pb_row.add_widget(self.lbl_pct)
        prog_box.add_widget(pb_row)
        root.add_widget(prog_box)

        # Kiritish maydoni
        inp_row = BoxLayout(orientation='horizontal', spacing=dp(6),
                            size_hint_y=None, height=dp(48))
        self.entry = TextInput(hint_text='Markirovka kodini kiriting...',
                               multiline=False, font_size=sp(13),
                               font_name='RobotoMono',
                               size_hint_x=0.75)
        self.entry.bind(on_text_validate=lambda *_: self._check_mark())
        btn_check = KivyBtn(text='Tekshir', bg=C_BLUE,
                            size_hint_x=0.25)
        btn_check.bind(on_press=lambda *_: self._check_mark())
        inp_row.add_widget(self.entry)
        inp_row.add_widget(btn_check)
        root.add_widget(inp_row)

        # Natija badge
        self.result_box = BoxLayout(orientation='vertical',
                                    size_hint_y=None, height=dp(60),
                                    padding=dp(10))
        with self.result_box.canvas.before:
            Color(*C_PANEL)
            self._res_rect = RoundedRectangle(pos=self.result_box.pos,
                                              size=self.result_box.size,
                                              radius=[dp(8)])
        self.result_box.bind(pos=self._upd_res, size=self._upd_res)
        self.lbl_result = Label(text='JSON faylni oching va skanerlang',
                                font_size=sp(13), bold=True,
                                color=C_MUTED, halign='left')
        self.lbl_result.bind(size=lambda w,_: setattr(w,'text_size',w.size))
        self.lbl_prod = Label(text='', font_size=sp(11),
                              color=C_MUTED, halign='left')
        self.lbl_prod.bind(size=lambda w,_: setattr(w,'text_size',w.size))
        self.result_box.add_widget(self.lbl_result)
        self.result_box.add_widget(self.lbl_prod)
        root.add_widget(self.result_box)

        # Eksport tugmalari
        exp_grid = GridLayout(cols=2, spacing=dp(6),
                              size_hint_y=None, height=dp(50))
        btn_pdf = KivyBtn(text='HTML hisobot', bg=C_RED)
        btn_pdf.bind(on_press=lambda *_: self._export('pdf'))
        btn_xl  = KivyBtn(text='Excel saqlash', bg=C_GREEN)
        btn_xl.bind(on_press=lambda *_: self._export('excel'))
        exp_grid.add_widget(btn_pdf)
        exp_grid.add_widget(btn_xl)
        root.add_widget(exp_grid)

        btn_clear = KivyBtn(text='Jurnalni tozalash',
                            bg=get_color_from_hex('#64748B'))
        btn_clear.bind(on_press=lambda *_: self._clear())
        root.add_widget(btn_clear)

        root.add_widget(Widget())  # spacer
        return root

    def _upd_res(self, *_):
        self._res_rect.pos  = self.result_box.pos
        self._res_rect.size = self.result_box.size

    # ── Jurnal tab ────────────────────────────────────────────────────────
    def _build_log_tab(self):
        self.log_layout = GridLayout(cols=1, spacing=dp(2),
                                     size_hint_y=None)
        self.log_layout.bind(minimum_height=self.log_layout.setter('height'))
        sv = ScrollView()
        sv.add_widget(self.log_layout)
        return sv

    # ── Yetishmayotganlar tab ─────────────────────────────────────────────
    def _build_missing_tab(self):
        self.miss_layout = GridLayout(cols=1, spacing=dp(2),
                                      size_hint_y=None)
        self.miss_layout.bind(minimum_height=self.miss_layout.setter('height'))
        sv = ScrollView()
        sv.add_widget(self.miss_layout)
        return sv

    # ── JSON ochish ────────────────────────────────────────────────────────
    def _open_json(self, *_):
        if ANDROID:
            request_permissions([Permission.READ_EXTERNAL_STORAGE,
                                  Permission.WRITE_EXTERNAL_STORAGE])

        content = BoxLayout(orientation='vertical', padding=dp(8), spacing=dp(6))
        start_path = _get_storage() if ANDROID else os.path.expanduser('~')
        fc = FileChooserListView(path=start_path,
                                 filters=['*.json'],
                                 size_hint_y=0.85)
        content.add_widget(fc)

        btn_row = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))
        btn_ok  = KivyBtn(text='Ochish',  bg=C_GREEN)
        btn_can = KivyBtn(text='Bekor',   bg=C_RED)
        btn_row.add_widget(btn_ok); btn_row.add_widget(btn_can)
        content.add_widget(btn_row)

        popup = Popup(title='JSON faktura faylini tanlang',
                      content=content, size_hint=(0.95, 0.9))

        def on_ok(*_):
            if not fc.selection:
                return
            popup.dismiss()
            self._load_json(fc.selection[0])

        btn_ok.bind(on_press=on_ok)
        btn_can.bind(on_press=popup.dismiss)
        popup.open()

    def _load_json(self, path):
        try:
            with open(path, encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            self._alert('Xato', f'Fayl o\'qishda xato:\n{e}')
            return

        fi = data.get('facturadoc', {})
        self.fi = {
            'no':   fi.get('facturano','?'),
            'date': fi.get('facturadate','?'),
            'seller_name':     data.get('seller',{}).get('name',''),
            'seller_tin':      data.get('sellertin',''),
            'seller_address':  data.get('seller',{}).get('address','').strip(),
            'seller_director': data.get('seller',{}).get('director',''),
            'buyer_name':      data.get('buyer',{}).get('name',''),
            'buyer_tin':       data.get('buyertin',''),
            'buyer_address':   data.get('buyer',{}).get('address','').strip(),
            'buyer_director':  data.get('buyer',{}).get('director',''),
        }
        self.products  = []
        self.all_marks = {}
        for p in data.get('productlist',{}).get('products',[]):
            prod = {'ordno':p['ordno'],'name':p['name'],
                    'count':p['count'],'marks':p.get('marks',{}).get('kiz',[])}
            self.products.append(prod)
            for m in prod['marks']:
                self.all_marks[m] = prod

        self.scanned_set.clear()
        self.scan_log.clear()
        self.stats = {'total':0,'ok':0,'err':0,'dup':0}
        total = sum(len(p['marks']) for p in self.products)
        self.lbl_title.text = f"Faktura №{self.fi['no']} | {total} ta markirovka"
        self._update_all()
        self._alert('Yuklandi',
                    f"Faktura №{self.fi['no']}\n{len(self.products)} mahsulot, {total} markirovka")

    # ── Markirovka tekshirish ─────────────────────────────────────────────
    def _check_mark(self):
        if not self.all_marks:
            self._alert('Diqqat', 'Avval JSON faktura faylini oching!')
            return
        val = self.entry.text.strip()
        if not val:
            return
        self.entry.text = ''

        if val in self.all_marks:
            prod = self.all_marks[val]
            if val in self.scanned_set:
                status = 'dup'
                msg    = "TAKRORIY — bu markirovka avval skanerlangan!"
                rb, rf = C_DUP_BG, C_DUP_FG
            else:
                status = 'ok'
                msg    = "TO'G'RI — markirovka topildi"
                rb, rf = C_OK_BG, C_OK_FG
                self.scanned_set.add(val)
            pname = f"#{prod['ordno']}  {prod['name'][:60]}"
        else:
            status = 'err'
            msg    = "XATO — bu markirovka fatura ro'yxatida yo'q!"
            rb, rf = C_ERR_BG, C_ERR_FG
            prod   = None
            pname  = 'Topilmadi'

        play_sound(status)
        self.stats['total'] += 1
        self.stats[status]  += 1
        self.scan_log.insert(0,{
            'mark':val,'status':status,
            'ordno':prod['ordno'] if prod else None,
            'prodname':prod['name'] if prod else 'Topilmadi'
        })

        # Badge yangilash
        with self.result_box.canvas.before:
            Color(*rb)
            self._res_rect = RoundedRectangle(pos=self.result_box.pos,
                                              size=self.result_box.size,
                                              radius=[dp(8)])
        self.lbl_result.color = rf
        self.lbl_result.text  = msg
        self.lbl_prod.color   = rf
        self.lbl_prod.text    = pname

        self._update_all()

    def _update_all(self):
        self._update_stats()
        self._refresh_log()
        self._refresh_missing()

    def _update_stats(self):
        self.card_total.set_value(self.stats['total'])
        self.card_ok.set_value(self.stats['ok'])
        self.card_err.set_value(self.stats['err'])
        self.card_dup.set_value(self.stats['dup'])
        total_m = sum(len(p['marks']) for p in self.products)
        scanned = len(self.scanned_set)
        remain  = max(0, total_m - scanned)
        pct     = int(scanned/total_m*100) if total_m > 0 else 0
        self.card_scanned.set_value(scanned)
        self.card_remain.set_value(remain)
        self.card_total_m.set_value(total_m)
        self.prog_bar.value = pct
        self.lbl_pct.text   = f'{pct}%'

    def _refresh_log(self):
        self.log_layout.clear_widgets()
        labels = {'ok':"To'g'ri",'err':'Xato','dup':'Takroriy'}
        colors_map = {'ok':C_OK_BG,'err':C_ERR_BG,'dup':C_DUP_BG}
        fg_map     = {'ok':C_OK_FG,'err':C_ERR_FG,'dup':C_DUP_FG}
        for r in self.scan_log:
            row = BoxLayout(orientation='horizontal', size_hint_y=None,
                            height=dp(44), padding=[dp(8),dp(4)], spacing=dp(6))
            with row.canvas.before:
                Color(*colors_map[r['status']])
                Rectangle(pos=row.pos, size=row.size)
            row.bind(pos=lambda w,_,rw=row: self._redraw_row(rw),
                     size=lambda w,_,rw=row: self._redraw_row(rw))

            lbl_st = Label(text=labels[r['status']], font_size=sp(11),
                           bold=True, color=fg_map[r['status']],
                           size_hint_x=0.18)
            lbl_m  = Label(text=r['mark'][:28]+'…' if len(r['mark'])>28 else r['mark'],
                           font_size=sp(10), color=fg_map[r['status']],
                           size_hint_x=0.45, halign='left')
            lbl_m.bind(size=lambda w,_: setattr(w,'text_size',w.size))
            lbl_p  = Label(text=r['prodname'][:25]+'…' if len(r['prodname'])>25 else r['prodname'],
                           font_size=sp(10), color=fg_map[r['status']],
                           size_hint_x=0.37, halign='left')
            lbl_p.bind(size=lambda w,_: setattr(w,'text_size',w.size))
            row.add_widget(lbl_st); row.add_widget(lbl_m); row.add_widget(lbl_p)
            self.log_layout.add_widget(row)

    def _redraw_row(self, row):
        row.canvas.before.clear()

    def _refresh_missing(self):
        self.miss_layout.clear_widgets()
        for p in self.products:
            for m in p['marks']:
                if m not in self.scanned_set:
                    row = BoxLayout(orientation='horizontal',
                                   size_hint_y=None, height=dp(40),
                                   padding=[dp(8),dp(4)], spacing=dp(6))
                    with row.canvas.before:
                        Color(*C_ERR_BG)
                        Rectangle(pos=row.pos, size=row.size)
                    lbl_o = Label(text=str(p['ordno']), font_size=sp(11),
                                  color=C_ERR_FG, size_hint_x=0.08)
                    lbl_m = Label(text=m[:30]+'…' if len(m)>30 else m,
                                  font_size=sp(10), color=C_ERR_FG,
                                  size_hint_x=0.48, halign='left')
                    lbl_m.bind(size=lambda w,_: setattr(w,'text_size',w.size))
                    lbl_n = Label(text=p['name'][:22]+'…',
                                  font_size=sp(10), color=C_ERR_FG,
                                  size_hint_x=0.44, halign='left')
                    lbl_n.bind(size=lambda w,_: setattr(w,'text_size',w.size))
                    row.add_widget(lbl_o); row.add_widget(lbl_m); row.add_widget(lbl_n)
                    self.miss_layout.add_widget(row)

    def _clear(self):
        self.scanned_set.clear()
        self.scan_log.clear()
        self.stats = {'total':0,'ok':0,'err':0,'dup':0}
        self.lbl_result.text  = 'Jurnal tozalandi'
        self.lbl_result.color = C_MUTED
        self.lbl_prod.text    = ''
        self._update_all()

    # ── Eksport ────────────────────────────────────────────────────────────
    def _export(self, fmt):
        if not self.products:
            self._alert('Diqqat', 'Avval JSON faktura faylini oching!')
            return
        ts    = datetime.now().strftime('%Y%m%d_%H%M%S')
        ext = 'html' if fmt == 'pdf' else 'xlsx'
        fname = f"markirovka_{self.fi.get('no','')}_hisobot_{ts}.{ext}"
        if ANDROID:
            save_dir = os.path.join(primary_external_storage_path(),
                                    'Download', 'MarkirovkaTizimi')
        else:
            save_dir = os.path.join(os.path.expanduser('~'), 'MarkirovkaTizimi')
        os.makedirs(save_dir, exist_ok=True)
        fpath = os.path.join(save_dir, fname)

        def _run():
            try:
                if fmt == 'pdf':
                    _make_pdf(fpath, self.products, self.scanned_set,
                              self.scan_log, self.stats, self.fi)
                else:
                    _make_excel(fpath, self.products, self.scanned_set,
                                self.scan_log, self.stats, self.fi)
                Clock.schedule_once(lambda _: self._alert(
                    'Tayyor!', f'Saqlandi:\n{fpath}'), 0)
            except Exception as e:
                Clock.schedule_once(lambda _,err=e: self._alert('Xato', str(err)), 0)

        threading.Thread(target=_run, daemon=True).start()

    def _alert(self, title, msg):
        content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
        content.add_widget(Label(text=msg, font_size=sp(13),
                                 halign='center', valign='middle'))
        btn = KivyBtn(text='OK', bg=C_BLUE, size_hint_y=None, height=dp(44))
        content.add_widget(btn)
        popup = Popup(title=title, content=content,
                      size_hint=(0.85, 0.4))
        btn.bind(on_press=popup.dismiss)
        popup.open()


# ══════════════════════════════════════════════════════════════════════════
#  APP
# ══════════════════════════════════════════════════════════════════════════
class MarkirovkaApp(App):
    def build(self):
        Window.clearcolor = (*C_BG[:3], 1)
        if ANDROID:
            request_permissions([Permission.READ_EXTERNAL_STORAGE,
                                  Permission.WRITE_EXTERNAL_STORAGE])
        return MainScreen()

    def get_application_name(self):
        return 'Markirovka Tizimi'


if __name__ == '__main__':
    MarkirovkaApp().run()
